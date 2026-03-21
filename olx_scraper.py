"""
Skrypt do scrapowania ofert wynajmu mieszkan w Gdansku.
Platformy: OLX, Otodom
Wymaga: pip install playwright openpyxl Pillow
         python -m playwright install chromium
"""

import csv
import io
import os
import re
import subprocess
import sys
import time
import urllib.request
from playwright.sync_api import sync_playwright

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── Konfiguracja ──────────────────────────────────────────────────────────────
BUDGET_LIMIT = 3500
MIN_AREA = 30
OUTPUT_CSV = "C:/Users/gabri/Desktop/mieszkania_gdansk.csv"
OUTPUT_XLSX = "C:/Users/gabri/Desktop/mieszkania_gdansk.xlsx"
IMAGES_DIR = "C:/Users/gabri/Desktop/mieszkania_zdjecia"

OLX_URL = "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/gdansk/"
OTODOM_URL = "https://www.otodom.pl/pl/wyniki/wynajem/mieszkanie/pomorskie/gdansk/gdansk/gdansk?limit=72&by=DEFAULT&direction=DESC"

# Slowa kluczowe
KW_ANIMALS_YES = [
    "zwierzęta akceptowane", "zwierzęta mile widziane", "akceptuję zwierzęta",
    "zwierzęta dozwolone", "można ze zwierzętami", "pet friendly", "zwierzęta tak",
    "akceptujemy zwierzęta", "zwierzaki ok", "zwierzaki mile widziane",
    "animals accepted", "pets allowed", "ze zwierzętami",
]
KW_ANIMALS_NO = [
    "bez zwierząt", "zwierzęta nieakceptowane", "nie akceptuję zwierząt",
    "zakaz trzymania zwierząt", "zwierzęta nie", "no pets",
    "nie akceptujemy zwierząt",
]
KW_BALCONY = ["balkon", "taras", "loggia", "balkony"]
KW_GYM = ["siłownia", "fitness", "sala gimnastyczna", "gym", "silownia"]
KW_GATED = ["osiedle strzeżone", "ochrona 24", "monitoring", "strzeżone", "gated", "ochrona"]
KW_WIFI = ["wifi", "wi-fi", "internet w cenie", "internet wliczony", "internet bezprzewodowy"]
KW_HEATING = [
    "ogrzewanie miejskie", "ogrzewanie gazowe", "ogrzewanie elektryczne",
    "co miejskie", "co gazowe", "piece gazowe", "pompa ciepła",
    "ogrzewanie podłogowe", "centralne ogrzewanie",
]
KW_HOT_WATER = [
    "ciepła woda miejska", "ciepła woda gazowa", "bojler", "podgrzewacz",
    "cwu", "ciepła woda elektryczna", "piec gazowy",
]
KW_INCOME_PROOF = [
    "potwierdzenie dochodu", "zaświadczenie o zarobkach", "weryfikacja dochodu",
    "potwierdzenie zarobków", "udokumentowane dochody", "umowa o pracę wymagana",
]


def find_keywords(text, keywords):
    t = text.lower()
    for kw in keywords:
        if kw.lower() in t:
            return kw
    return None


def extract_number(text, pattern):
    m = re.search(pattern, text, re.IGNORECASE)
    if m:
        return m.group(1).replace(" ", "").replace(",", ".")
    return ""


def dismiss_cookies(page):
    selectors = [
        "button#onetrust-accept-btn-handler",
        "button:has-text('Akceptuję')",
        "button:has-text('Zgadzam się')",
        "button:has-text('Zaakceptuj')",
        "button:has-text('Accept')",
        "[data-testid='gdpr-consent-accept']",
        "button:has-text('Akceptuj')",
        "button:has-text('Zgoda')",
    ]
    for sel in selectors:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=1500):
                btn.click()
                print("  [OK] Zamknieto cookie banner")
                time.sleep(1)
                return
        except Exception:
            continue


def download_image(url, filepath):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            with open(filepath, "wb") as f:
                f.write(resp.read())
        return True
    except Exception:
        return False


def parse_price_value(s):
    """Parsuje string z cena na float."""
    if not s or s == "Brak danych":
        return None
    s = s.replace(" ", "").replace(",", ".").replace("zł", "").replace("zl", "").replace("/mies", "").replace("/mc", "")
    nums = re.findall(r"[\d.]+", s)
    return float(nums[0]) if nums else None


def extract_costs_from_text(text):
    """Skanuje opis szukajac kosztow. Pracuje na liniach zeby unikac falszywych dopasowań."""
    costs = {
        "czynsz_admin": None, "media": None, "prad": None, "gaz": None,
        "woda": None, "ogrzewanie_koszt": None, "internet": None, "tv": None,
        "smieci": None, "parking": None, "ubezpieczenie": None,
        "kaucja": None, "kaucja_opis": None, "prowizja": None,
        "inne_oplaty": [], "uwagi_kosztowe": [],
    }

    t = text.lower()
    lines = t.split("\n")

    # Szukaj kosztow linia po linii - mniejsze ryzyko falszywych dopasowań
    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Czynsz administracyjny / oplaty
        if any(k in line for k in ["czynsz admin", "czynsz eksploat", "opłat", "oplat", "do spółdzielni", "do wspolnoty", "do wspólnoty", "zaliczka"]):
            if not any(k in line for k in ["czynsz najmu", "najem"]):
                m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
                if m and costs["czynsz_admin"] is None:
                    val = float(m.group(1).replace(" ", "").replace(",", "."))
                    if 50 <= val <= 2000:  # rozsadny zakres
                        costs["czynsz_admin"] = val

        # Media (zbiorczo)
        if "media" in line and "w cenie" not in line:
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["media"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 50 <= val <= 1500:
                    costs["media"] = val

        # Prad
        if any(k in line for k in ["prąd", "prad", "energia elektr"]):
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["prad"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 30 <= val <= 800:
                    costs["prad"] = val

        # Gaz
        if "gaz" in line and "garaż" not in line and "garaz" not in line:
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["gaz"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 20 <= val <= 500:
                    costs["gaz"] = val

        # Woda
        if "woda" in line:
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["woda"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 20 <= val <= 400:
                    costs["woda"] = val

        # Ogrzewanie (koszt)
        if any(k in line for k in ["ogrzewanie", "c.o.", "co ", "centralne ogrzew"]):
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["ogrzewanie_koszt"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 30 <= val <= 800:
                    costs["ogrzewanie_koszt"] = val

        # Internet
        if any(k in line for k in ["internet", "wifi", "wi-fi"]) and "w cenie" not in line:
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["internet"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 20 <= val <= 200:
                    costs["internet"] = val

        # Smieci
        if any(k in line for k in ["śmieci", "smieci", "odpady"]):
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["smieci"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 10 <= val <= 200:
                    costs["smieci"] = val

        # Parking
        if any(k in line for k in ["parking", "garaż", "garaz", "miejsce parkingowe", "miejsce postojowe"]):
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["parking"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if 50 <= val <= 800:
                    costs["parking"] = val

        # Kaucja
        if any(k in line for k in ["kaucj", "depozyt"]):
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m and costs["kaucja"] is None:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                if val >= 500:
                    costs["kaucja"] = val
            # Kaucja jako wielokrotnosc
            m2 = re.search(r"(\d)\s*(?:x|×|\*|krotno)", line)
            if m2 and costs["kaucja"] is None:
                costs["kaucja_opis"] = f"{m2.group(1)}x czynsz"

        # Prowizja
        if "prowizj" in line or "pośrednictw" in line or "posrednictw" in line:
            m = re.search(r"(\d[\d\s]*(?:[,.]\d+)?)\s*(?:z[lł]|pln)", line)
            if m:
                val = float(m.group(1).replace(" ", "").replace(",", "."))
                costs["prowizja"] = val

    # Uwagi kosztowe (szukaj w calym tekscie)
    if "media w cenie" in t or "media wliczone" in t or "zawiera media" in t:
        costs["uwagi_kosztowe"].append("media w cenie czynszu")
    if "internet w cenie" in t or "internet wliczony" in t:
        costs["uwagi_kosztowe"].append("internet w cenie")
    if "parking w cenie" in t or "miejsce parkingowe w cenie" in t:
        costs["uwagi_kosztowe"].append("parking w cenie")
    if re.search(r"bez\s*prowizji|0\s*%\s*prowizji|prowizja\s*0", t):
        costs["prowizja"] = 0
        costs["uwagi_kosztowe"].append("bez prowizji")

    return costs


def get_title(page):
    """Probuje wyciagnac tytul ogloszenia roznymi selektorami."""
    selectors = [
        "h1",
        "[data-cy='ad_title']",
        "[data-testid='ad-title']",
        "[data-testid='ad_title']",
        "h1[data-cy='ad_title']",
        "h4.css-1juynto",         # OLX nowy layout
        ".css-1juynto",
        "[class*='title'] h1",
        "[class*='Title']",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1500):
                txt = el.inner_text().strip()
                if txt and len(txt) > 3:
                    return txt
        except Exception:
            continue

    # Fallback: tytul z tagu <title>
    try:
        title_tag = page.title()
        if title_tag:
            # OLX: "Tytul ogloszenia - OLX.pl"
            clean = re.sub(r'\s*[-|–]\s*(OLX|Otodom).*$', '', title_tag).strip()
            if clean and len(clean) > 3:
                return clean
    except Exception:
        pass

    # Fallback: wyciagnij nazwe z URL
    try:
        url = page.url
        # np. /d/oferta/wynajme-mieszkanie-2-pokojowe-gdansk-CID3-ID...
        m = re.search(r'/oferta/([^?]+?)(?:-CID\d|-ID\w)', url)
        if m:
            name_from_url = m.group(1).replace("-", " ").strip().title()
            if len(name_from_url) > 5:
                return name_from_url
    except Exception:
        pass

    return "Brak danych"


def get_photos(page):
    """Wyciaga URLe zdjec z ogloszenia."""
    img_urls = []

    # Selektory zdjec (OLX + Otodom)
    img_selectors = [
        "div[data-testid='ad-photo'] img",
        "[data-testid='ad-image'] img",
        ".swiper-slide img",
        "[class*='photo'] img",
        "[class*='gallery'] img",
        "[class*='image-gallery'] img",
        "[data-cy='gallery'] img",
        "picture img",
        "[class*='slider'] img",
    ]
    for sel in img_selectors:
        try:
            imgs = page.locator(sel).all()
            for img in imgs:
                src = img.get_attribute("src") or img.get_attribute("data-src") or ""
                if src and src.startswith("http") and any(d in src for d in ["olxcdn", "otodom", "ireland.apollo"]):
                    if src not in img_urls:
                        img_urls.append(src)
        except Exception:
            continue

    # Fallback: szukaj w HTML
    if not img_urls:
        try:
            html = page.content()
            found = re.findall(r'https://[^"\']+(?:olxcdn|otodom)[^"\']+\.(?:jpg|jpeg|webp|png)', html)
            for u in found:
                if u not in img_urls:
                    img_urls.append(u)
        except Exception:
            pass

    return img_urls[:10]


def extract_listing_data(page, url, listing_index, platform="OLX"):
    """Wchodzi w ogloszenie i wyciaga dane."""
    data = {
        "Platforma": platform,
        "URL": url,
        "Nazwa": "Brak danych",
        "Cena_Czynsz_Najmu": "Brak danych",
        "Czynsz_Administracyjny": "Brak danych",
        "Media": "Brak danych",
        "Prad": "Brak danych",
        "Gaz": "Brak danych",
        "Woda": "Brak danych",
        "Internet": "Brak danych",
        "Smieci": "Brak danych",
        "Parking": "Brak danych",
        "Prowizja": "Brak danych",
        "Cena_Suma": "Brak danych",
        "Koszty_Rozpisane": "",
        "Powierzchnia": "Brak danych",
        "Kaucja": "Brak danych",
        "Termin": "Brak danych",
        "Min_Dlugosc_Najmu": "Brak danych",
        "Potwierdzenie_Dochodu": "Brak danych",
        "Zwierzeta": "Brak danych",
        "Balkon": "Brak danych",
        "Silownia": "Brak danych",
        "Osiedle_Strzezone": "Brak danych",
        "WiFi": "Brak danych",
        "Ogrzewanie": "Brak danych",
        "Ciepla_Woda": "Brak danych",
        "Data_Wystawienia": "Brak danych",
        "Uwagi": "",
        "Zdjecia_URL": [],
        "Zdjecie_Lokalne": "",
    }

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)
        dismiss_cookies(page)

        # ── Nazwa ──
        data["Nazwa"] = get_title(page)

        # ── Data wystawienia ──
        date_selectors = [
            "[data-cy='ad-posted-at']",
            "[data-testid='ad-posted-at']",
            "span[class*='date']",
            "[class*='creation-date']",
            "[class*='DateTimeStamp']",
        ]
        for sel in date_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible(timeout=1500):
                    data["Data_Wystawienia"] = el.inner_text().strip()
                    break
            except Exception:
                continue
        if data["Data_Wystawienia"] == "Brak danych":
            # Szukaj w tekscie: "Dodane o HH:MM, DD miesiąc RRRR"
            m = re.search(
                r"(?:dodane?|opublikowane?|wystawione?)[^a-z]{0,10}(.*?\d{4}|\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4}|dzisiaj|wczoraj|\d+\s*(?:min|godz|dni)\s*temu)",
                page.inner_text("body"), re.IGNORECASE,
            )
            if m:
                data["Data_Wystawienia"] = m.group(0).strip()

        # ── Zdjecia ──
        img_urls = get_photos(page)
        data["Zdjecia_URL"] = img_urls
        if img_urls:
            os.makedirs(IMAGES_DIR, exist_ok=True)
            img_path = os.path.join(IMAGES_DIR, f"mieszkanie_{listing_index:03d}.jpg")
            if download_image(img_urls[0], img_path):
                data["Zdjecie_Lokalne"] = img_path
                print(f"  [img] Pobrano ({len(img_urls)} zdjec)")

        # ── Tekst strony ──
        full_text = page.inner_text("body")

        # ── Cena (czynsz najmu) ──
        price_selectors = [
            "[data-testid='ad-price-container'] h3",
            "[data-cy='ad_price']",
            "[data-testid='ad-price']",
            "[aria-label='Cena']",
            "h3[class*='price']",
        ]
        for sel in price_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible(timeout=1500):
                    txt = el.inner_text().strip()
                    if re.search(r'\d', txt):
                        data["Cena_Czynsz_Najmu"] = txt
                        break
            except Exception:
                continue

        if data["Cena_Czynsz_Najmu"] == "Brak danych":
            # Szukaj ceny w pierwszych liniach
            for line in full_text.split("\n")[:30]:
                m = re.search(r"(\d[\d\s]*(?:,\d+)?)\s*zł\s*/\s*mies", line, re.IGNORECASE)
                if m:
                    data["Cena_Czynsz_Najmu"] = m.group(0).strip()
                    break

        # ── Powierzchnia ──
        # Otodom ma parametry w tabeli
        area_patterns = [
            r"(?:powierzchnia|pow\.?)[:\s]*(\d[\d\s]*(?:[,\.]\d+)?)\s*m",
            r"(\d{2,3}(?:[,\.]\d+)?)\s*m[²2]",
        ]
        for pat in area_patterns:
            area = extract_number(full_text, pat)
            if area:
                data["Powierzchnia"] = area + " m2"
                break

        # ── Dokladna analiza kosztow ──
        costs = extract_costs_from_text(full_text)

        if costs["czynsz_admin"]:
            data["Czynsz_Administracyjny"] = f"{costs['czynsz_admin']:.0f} zl"
        if costs["media"]:
            data["Media"] = f"{costs['media']:.0f} zl"
        if costs["prad"]:
            data["Prad"] = f"{costs['prad']:.0f} zl"
        if costs["gaz"]:
            data["Gaz"] = f"{costs['gaz']:.0f} zl"
        if costs["woda"]:
            data["Woda"] = f"{costs['woda']:.0f} zl"
        if costs["internet"]:
            data["Internet"] = f"{costs['internet']:.0f} zl"
        if costs["smieci"]:
            data["Smieci"] = f"{costs['smieci']:.0f} zl"
        if costs["parking"]:
            data["Parking"] = f"{costs['parking']:.0f} zl"
        if costs["prowizja"] is not None:
            data["Prowizja"] = f"{costs['prowizja']:.0f} zl" if costs["prowizja"] > 0 else "Brak"

        if costs["kaucja"]:
            data["Kaucja"] = f"{costs['kaucja']:.0f} zl"
        elif costs.get("kaucja_opis"):
            data["Kaucja"] = costs["kaucja_opis"]

        # ── Termin ──
        term_match = re.search(
            r"(?:dost[eę]pn[eay]|od\s+dnia|termin|wprowadz)[^a-z]{0,10}(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4}|od\s*zaraz|natychmiast)",
            full_text, re.IGNORECASE,
        )
        if term_match:
            data["Termin"] = term_match.group(0).strip()
        elif re.search(r"od\s*zaraz|natychmiast", full_text, re.IGNORECASE):
            data["Termin"] = "Od zaraz"

        # Min. dlugosc najmu
        min_rent = re.search(
            r"(?:min(?:imaln[ay])?\.?\s*(?:okres|czas|d[lł]ugo[sś][cć])?\s*(?:najmu|wynajmu)?)[^0-9]*(\d+)\s*(mies|rok|lat)",
            full_text, re.IGNORECASE,
        )
        if min_rent:
            data["Min_Dlugosc_Najmu"] = f"{min_rent.group(1)} {min_rent.group(2)}"

        # ── Slowa kluczowe ──
        if find_keywords(full_text, KW_ANIMALS_YES):
            data["Zwierzeta"] = "Tak"
        elif find_keywords(full_text, KW_ANIMALS_NO):
            data["Zwierzeta"] = "Nie"

        if find_keywords(full_text, KW_BALCONY):
            data["Balkon"] = "Tak"
        if find_keywords(full_text, KW_GYM):
            data["Silownia"] = "Tak"
        if find_keywords(full_text, KW_GATED):
            data["Osiedle_Strzezone"] = "Tak"
        if find_keywords(full_text, KW_WIFI):
            data["WiFi"] = "Tak"

        heat = find_keywords(full_text, KW_HEATING)
        if heat:
            data["Ogrzewanie"] = heat.capitalize()
        hw = find_keywords(full_text, KW_HOT_WATER)
        if hw:
            data["Ciepla_Woda"] = hw.capitalize()
        if find_keywords(full_text, KW_INCOME_PROOF):
            data["Potwierdzenie_Dochodu"] = "Tak"

        # ── Oblicz REALNA cene ──
        rent = parse_price_value(data["Cena_Czynsz_Najmu"])

        total = 0.0
        cost_breakdown = []

        if rent is not None:
            total += rent
            cost_breakdown.append(f"najem: {rent:.0f}")

        for label, key in [
            ("admin", "czynsz_admin"), ("media", "media"), ("prad", "prad"),
            ("gaz", "gaz"), ("woda", "woda"), ("ogrzewanie", "ogrzewanie_koszt"),
            ("internet", "internet"), ("tv", "tv"), ("smieci", "smieci"),
            ("parking", "parking"), ("ubezpieczenie", "ubezpieczenie"),
        ]:
            val = costs.get(key)
            if val:
                total += val
                cost_breakdown.append(f"{label}: {val:.0f}")

        for label, val in costs.get("inne_oplaty", []):
            total += val
            cost_breakdown.append(f"{label}: {val:.0f}")

        media_found = any(costs.get(k) for k in ["media", "prad", "gaz", "woda", "ogrzewanie_koszt"])
        if not media_found and rent is not None:
            total += 250
            cost_breakdown.append("media ~250 (szacunek)")

        if rent is not None:
            data["Cena_Suma"] = f"{total:.0f} zl"
            data["Koszty_Rozpisane"] = " + ".join(cost_breakdown)
            if total > BUDGET_LIMIT:
                data["Uwagi"] = "POZA BUDZETEM"

        if costs["uwagi_kosztowe"]:
            notes = "; ".join(costs["uwagi_kosztowe"])
            data["Uwagi"] += ("; " if data["Uwagi"] else "") + notes

        if data["Powierzchnia"] != "Brak danych":
            area_val = parse_price_value(data["Powierzchnia"])
            if area_val and area_val < MIN_AREA:
                data["Uwagi"] += ("; " if data["Uwagi"] else "") + f"MALE (<{MIN_AREA}m2)"

    except Exception as e:
        data["Uwagi"] = f"BLAD: {e}"

    return data


# ── Zbieranie linkow ──────────────────────────────────────────────────────────

def collect_olx_links(page, max_pages=5):
    """Zbiera linki z OLX."""
    all_links = set()
    for page_num in range(1, max_pages + 1):
        url = OLX_URL if page_num == 1 else f"{OLX_URL}?page={page_num}"
        print(f"  [OLX] Strona {page_num}...")
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except Exception:
            pass
        time.sleep(3)
        if page_num == 1:
            dismiss_cookies(page)

        anchors = page.locator("a[href*='/oferta/'], a[href*='/d/oferta/']").all()
        page_links = set()
        for a in anchors:
            try:
                href = a.get_attribute("href")
                if href and "/oferta/" in href:
                    if href.startswith("/"):
                        href = "https://www.olx.pl" + href
                    if "olx.pl" in href:
                        page_links.add(href)
            except Exception:
                continue
        print(f"       -> {len(page_links)} ofert")
        if not page_links:
            break
        all_links.update(page_links)

        try:
            next_btn = page.locator("[data-testid='pagination-forward']").first
            if not next_btn.is_visible(timeout=2000):
                break
        except Exception:
            break

    return sorted(all_links)


def collect_otodom_links(page, max_pages=3):
    """Zbiera linki z Otodom."""
    all_links = set()
    for page_num in range(1, max_pages + 1):
        url = OTODOM_URL if page_num == 1 else f"{OTODOM_URL}&page={page_num}"
        print(f"  [Otodom] Strona {page_num}...")
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except Exception:
            pass
        time.sleep(3)
        if page_num == 1:
            dismiss_cookies(page)

        # Otodom uzywa linkow /pl/oferta/...
        anchors = page.locator("a[href*='/pl/oferta/']").all()
        page_links = set()
        for a in anchors:
            try:
                href = a.get_attribute("href")
                if href and "/pl/oferta/" in href:
                    if href.startswith("/"):
                        href = "https://www.otodom.pl" + href
                    if "otodom.pl" in href:
                        page_links.add(href)
            except Exception:
                continue
        print(f"       -> {len(page_links)} ofert")
        if not page_links:
            break
        all_links.update(page_links)

        # Paginacja Otodom
        try:
            next_btn = page.locator("[data-cy='pagination.next-page'], button[aria-label='następna strona'], li.css-43nhzf a").first
            if not next_btn.is_visible(timeout=2000):
                break
        except Exception:
            break

    return sorted(all_links)


# ── Zapis wynikow ─────────────────────────────────────────────────────────────

def save_csv(results, filename):
    columns = [
        "Platforma", "Nazwa", "URL", "Zdjecie_1_URL", "Zdjecie_2_URL", "Zdjecie_3_URL",
        "Cena_Czynsz_Najmu", "Czynsz_Administracyjny",
        "Media", "Prad", "Gaz", "Woda", "Internet", "Smieci", "Parking", "Prowizja",
        "Cena_Suma", "Koszty_Rozpisane", "Powierzchnia", "Kaucja", "Termin",
        "Min_Dlugosc_Najmu", "Potwierdzenie_Dochodu", "Zwierzeta",
        "Balkon", "Silownia", "Osiedle_Strzezone", "WiFi",
        "Ogrzewanie", "Ciepla_Woda", "Data_Wystawienia", "Uwagi",
    ]
    rows = []
    for r in results:
        row = dict(r)
        photos = r.get("Zdjecia_URL", [])
        row["Zdjecie_1_URL"] = photos[0] if len(photos) > 0 else ""
        row["Zdjecie_2_URL"] = photos[1] if len(photos) > 1 else ""
        row["Zdjecie_3_URL"] = photos[2] if len(photos) > 2 else ""
        rows.append(row)

    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, delimiter=";", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"\n[OK] Zapisano CSV: {filename}")


def save_xlsx(results, filename):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Osobny arkusz per platforma + zbiorczy
    platforms = {}
    for r in results:
        p = r.get("Platforma", "Inne")
        platforms.setdefault(p, []).append(r)

    sheets_to_create = [("Wszystkie", results)] + [(p, items) for p, items in platforms.items()]

    columns = [
        "Zdjecie", "Platforma", "Nazwa", "URL",
        "Cena_Czynsz_Najmu", "Czynsz_Administracyjny",
        "Media", "Prad", "Gaz", "Woda", "Internet", "Smieci", "Parking", "Prowizja",
        "Cena_Suma", "Koszty_Rozpisane", "Powierzchnia", "Kaucja", "Termin",
        "Min_Dlugosc_Najmu", "Potwierdzenie_Dochodu", "Zwierzeta",
        "Balkon", "Silownia", "Osiedle_Strzezone", "WiFi",
        "Ogrzewanie", "Ciepla_Woda", "Data_Wystawienia", "Uwagi",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    link_font = Font(color="0563C1", underline="single")
    olx_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    otodom_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for sheet_idx, (sheet_name, sheet_data) in enumerate(sheets_to_create):
        if sheet_idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        # Naglowki
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 22

        for row_idx, item in enumerate(sheet_data, 2):
            ws.row_dimensions[row_idx].height = 90

            # Kolor tla wg platformy
            platform_fill = olx_fill if item.get("Platforma") == "OLX" else otodom_fill

            for col_idx, col_name in enumerate(columns, 1):
                if col_name == "Zdjecie":
                    img_path = item.get("Zdjecie_Lokalne", "")
                    if img_path and os.path.exists(img_path):
                        try:
                            img = XlImage(img_path)
                            img.width = 150
                            img.height = 112
                            ws.add_image(img, f"A{row_idx}")
                        except Exception:
                            ws.cell(row=row_idx, column=col_idx, value="(blad img)")
                    else:
                        ws.cell(row=row_idx, column=col_idx, value="Brak zdjecia")
                    continue

                if col_name == "URL":
                    url_val = item.get("URL", "")
                    cell = ws.cell(row=row_idx, column=col_idx, value="LINK")
                    cell.hyperlink = url_val
                    cell.font = link_font
                    continue

                val = item.get(col_name, "Brak danych")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)

                # Koloruj platforme
                if col_name == "Platforma":
                    cell.fill = platform_fill

                if col_name == "Powierzchnia" and val != "Brak danych":
                    nums = re.findall(r"[\d.]+", val.replace(",", "."))
                    if nums and float(nums[0]) < MIN_AREA:
                        cell.fill = red_fill
                        cell.font = red_font

                if col_name == "Uwagi" and "POZA BUDZETEM" in str(val):
                    cell.fill = red_fill
                    cell.font = red_font

        # Szerokosci kolumn
        for col_idx, col_name in enumerate(columns, 1):
            if col_name != "Zdjecie":
                ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 15)

    # Arkusz z linkami do zdjec
    ws2 = wb.create_sheet("Zdjecia - linki")
    ws2.cell(row=1, column=1, value="Platforma").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Nazwa").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="URL oferty").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    for i in range(1, 11):
        c = ws2.cell(row=1, column=i + 3, value=f"Zdjecie_{i}")
        c.font = header_font
        c.fill = header_fill

    for row_idx, item in enumerate(results, 2):
        ws2.cell(row=row_idx, column=1, value=item.get("Platforma", ""))
        ws2.cell(row=row_idx, column=2, value=item.get("Nazwa", ""))
        url_cell = ws2.cell(row=row_idx, column=3, value="LINK")
        url_cell.hyperlink = item.get("URL", "")
        url_cell.font = link_font
        for i, photo_url in enumerate(item.get("Zdjecia_URL", [])[:10]):
            c = ws2.cell(row=row_idx, column=i + 4, value="FOTO")
            c.hyperlink = photo_url
            c.font = link_font

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 15

    try:
        wb.save(filename)
        print(f"[OK] Zapisano XLSX: {filename}")
    except PermissionError:
        # Plik zablokowany — usun i sprobuj ponownie, lub zapisz pod inna nazwa
        try:
            os.remove(filename)
            wb.save(filename)
            print(f"[OK] Zapisano XLSX: {filename}")
        except Exception:
            alt = filename.replace(".xlsx", f"_{int(time.time())}.xlsx")
            wb.save(alt)
            print(f"[OK] Zapisano XLSX pod inna nazwa: {alt}")


def print_top3(results):
    scored = []
    for r in results:
        area = parse_price_value(r.get("Powierzchnia", "0")) or 0
        total = parse_price_value(r.get("Cena_Suma", "99999")) or 99999
        animals = r.get("Zwierzeta", "Brak danych").lower()

        score = 0
        if area >= MIN_AREA:
            score += 3
        if total <= BUDGET_LIMIT:
            score += 3
        if animals == "tak":
            score += 2
        if r.get("Balkon", "Brak danych") == "Tak":
            score += 1
        if r.get("Silownia", "Brak danych") == "Tak":
            score += 1
        if area > 0:
            score += max(0, 3 - (total / area / 30))
        scored.append((score, r))

    scored.sort(key=lambda x: -x[0])
    top3 = scored[:3]

    print("\n" + "=" * 80)
    print("  TOP 3 OFERT (metraz >30m2, cena <3500 zl, akceptacja zwierzat)")
    print("=" * 80)
    for i, (score, r) in enumerate(top3, 1):
        print(f"\n  #{i}: [{r['Platforma']}] {r['Nazwa']}")
        print(f"      Powierzchnia: {r['Powierzchnia']}")
        print(f"      Cena suma:    {r['Cena_Suma']}")
        print(f"      Zwierzeta:    {r['Zwierzeta']}")
        print(f"      Balkon:       {r['Balkon']}")
        print(f"      URL:          {r['URL']}")
    print("=" * 80)


def main():
    CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    CDP_PORT = 9222

    print("\n" + "=" * 60)
    print("  Skrypt przeszuka oferty wynajmu mieszkan w Gdansku")
    print("  Platformy: OLX, Otodom")
    print("=" * 60)
    print("\n[*] Zamykam Chrome...")
    os.system("taskkill /F /IM chrome.exe >nul 2>&1")
    time.sleep(2)

    print("[*] Uruchamiam Chrome...")
    subprocess.Popen([
        CHROME_PATH,
        f"--remote-debugging-port={CDP_PORT}",
        "--user-data-dir=C:\\Users\\gabri\\chrome-olx-debug",
        "--no-first-run",
        "--no-default-browser-check",
        "about:blank",
    ])
    time.sleep(5)

    try:
        urllib.request.urlopen(f"http://127.0.0.1:{CDP_PORT}/json/version", timeout=5)
        print("[OK] Chrome uruchomiony!")
    except Exception:
        print("[BLAD] Chrome nie nasluchuje na porcie 9222.")
        return

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{CDP_PORT}")
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = context.pages[0] if context.pages else context.new_page()
        time.sleep(2)

        all_results = []
        listing_counter = 0

        # ── OLX ──
        print("\n" + "=" * 40)
        print("  SKANOWANIE: OLX.pl")
        print("=" * 40)
        olx_links = collect_olx_links(page, max_pages=5)
        print(f"\n  -> OLX: {len(olx_links)} ofert")

        for i, link in enumerate(olx_links, 1):
            listing_counter += 1
            print(f"\n[OLX] [{i}/{len(olx_links)}] {link[:70]}...")
            data = extract_listing_data(page, link, listing_counter, "OLX")
            all_results.append(data)
            print(f"  -> {data['Nazwa'][:45]} | {data['Cena_Suma']} | {data['Powierzchnia']}")

        # ── Otodom ──
        print("\n" + "=" * 40)
        print("  SKANOWANIE: Otodom.pl")
        print("=" * 40)
        otodom_links = collect_otodom_links(page, max_pages=3)
        print(f"\n  -> Otodom: {len(otodom_links)} ofert")

        for i, link in enumerate(otodom_links, 1):
            listing_counter += 1
            print(f"\n[Otodom] [{i}/{len(otodom_links)}] {link[:70]}...")
            data = extract_listing_data(page, link, listing_counter, "Otodom")
            all_results.append(data)
            print(f"  -> {data['Nazwa'][:45]} | {data['Cena_Suma']} | {data['Powierzchnia']}")

        # ── Zapis ──
        print(f"\n{'=' * 60}")
        print(f"  PODSUMOWANIE: {len(all_results)} ofert")
        print(f"    OLX:    {len(olx_links)}")
        print(f"    Otodom: {len(otodom_links)}")
        print(f"{'=' * 60}")

        save_csv(all_results, OUTPUT_CSV)
        save_xlsx(all_results, OUTPUT_XLSX)
        print_top3(all_results)

        print(f"\n[OK] Pliki na Pulpicie:")
        print(f"   - {OUTPUT_CSV}")
        print(f"   - {OUTPUT_XLSX}")
        print(f"   - {IMAGES_DIR}/")

        print("\n[OK] Skonczono!")
        input("\n>>> Nacisnij ENTER aby zakonczyc... ")


if __name__ == "__main__":
    main()
